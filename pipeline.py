"""
pipeline.py
Automated Supply Chain Reporting Pipeline

Consolidates multiple raw Excel/CSV planner exports (each from a different
regional team, each with its own formatting quirks) into a single cleaned
dataset, then auto-generates a summary pivot report -- replacing the manual
weekly copy/paste/clean routine an analyst would otherwise do by hand.

Usage:
    python pipeline.py

Reads every file in data/raw_exports/ and writes:
    output/consolidated_data.csv
    output/Weekly_Supply_Chain_Summary_Report.xlsx
    output/pipeline_log.txt
"""
import glob
import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

RAW_DIR = "data/raw_exports"
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Canonical column names every file gets mapped onto, regardless of how the
# source file spelled/spaced/cased them.
COLUMN_ALIASES = {
    "sku": "SKU",
    "region": "Region",
    "date": "Date",
    "units sold": "Units Sold",
    "units_sold": "Units Sold",
}

log_lines = []

def log(msg):
    print(msg)
    log_lines.append(msg)


def normalize_columns(df):
    """Map any spelling/spacing/case variant of a column onto its canonical name."""
    new_cols = {}
    for col in df.columns:
        key = str(col).strip().lower().replace("_", " ")
        new_cols[col] = COLUMN_ALIASES.get(key, col.strip() if isinstance(col, str) else col)
    return df.rename(columns=new_cols)


def find_header_row(path, max_scan=5):
    """Some exports have metadata rows above the real header. Scan the first
    few lines of a CSV and return how many rows to skip before the header."""
    with open(path, "r", errors="ignore") as f:
        for i, line in enumerate(f):
            if i >= max_scan:
                break
            if "sku" in line.lower():
                return i
    return 0


def read_raw_file(path):
    """Read a single raw export, whatever its format, and return a DataFrame
    with normalized column names. Returns None (and logs why) if unusable."""
    fname = os.path.basename(path)
    ext = os.path.splitext(path)[1].lower()

    try:
        if ext == ".csv":
            skiprows = find_header_row(path)
            df = pd.read_csv(path, skiprows=skiprows)
        elif ext in (".xlsx", ".xls"):
            df = pd.read_excel(path)
        else:
            log(f"  SKIPPED {fname}: unsupported file type '{ext}'")
            return None
    except Exception as e:
        log(f"  SKIPPED {fname}: could not read file ({e})")
        return None

    df = normalize_columns(df)
    required = {"SKU", "Region", "Date", "Units Sold"}
    missing_cols = required - set(df.columns)
    if missing_cols:
        log(f"  SKIPPED {fname}: missing required column(s) {missing_cols}")
        return None

    df = df[["SKU", "Region", "Date", "Units Sold"]].copy()
    df["Source File"] = fname
    return df


def clean_dataset(df):
    """Apply all cleaning rules to the consolidated dataset and log what
    each rule removed or fixed, so the automation's impact is visible."""
    start_n = len(df)

    # Drop rows that are entirely blank (stray blank rows in some exports)
    df = df.dropna(how="all")

    # Standardize text fields
    df["SKU"] = df["SKU"].astype(str).str.strip().str.upper()
    df["Region"] = df["Region"].astype(str).str.strip().str.title()

    # Parse dates regardless of source format (MM/DD/YYYY, YYYY-MM-DD, or a
    # real Excel date already parsed by pandas). format="mixed" is required
    # here: once files with different date formats are concatenated into one
    # column, pandas' default fast-path infers a single format from the
    # first value and silently fails (returns NaT) on every row that
    # doesn't match it -- format="mixed" parses each value independently.
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce", format="mixed")
    n_bad_dates = df["Date"].isna().sum()
    if n_bad_dates:
        log(f"  Dropped {n_bad_dates} row(s) with an unparseable date")
    df = df.dropna(subset=["Date"])

    # Coerce Units Sold to numeric; anything non-numeric becomes NaN
    df["Units Sold"] = pd.to_numeric(df["Units Sold"], errors="coerce")

    n_missing_units = df["Units Sold"].isna().sum()
    if n_missing_units:
        log(f"  Dropped {n_missing_units} row(s) with missing Units Sold "
            f"(demand can't be safely assumed, so these are excluded rather than filled)")
    df = df.dropna(subset=["Units Sold"])

    n_negative = (df["Units Sold"] < 0).sum()
    if n_negative:
        log(f"  Dropped {n_negative} row(s) with a negative Units Sold value (data-entry error)")
    df = df[df["Units Sold"] >= 0]

    df["Units Sold"] = df["Units Sold"].astype(int)

    n_before_dupes = len(df)
    df = df.drop_duplicates(subset=["SKU", "Region", "Date", "Units Sold"])
    n_dupes = n_before_dupes - len(df)
    if n_dupes:
        log(f"  Dropped {n_dupes} exact duplicate row(s)")

    df = df.sort_values(["Date", "Region", "SKU"]).reset_index(drop=True)

    log(f"  Result: {start_n} raw rows -> {len(df)} clean rows")
    return df


def build_summary_report(df, out_path):
    """Auto-generate a multi-sheet Excel summary report from the cleaned data."""
    df["Week Starting"] = df["Date"].dt.to_period("W-SUN").apply(lambda p: p.start_time)

    sku_region_pivot = pd.pivot_table(
        df, index="SKU", columns="Region", values="Units Sold",
        aggfunc="sum", fill_value=0, margins=True, margins_name="Grand Total"
    )
    weekly_pivot = pd.pivot_table(
        df, index="SKU", columns="Week Starting", values="Units Sold",
        aggfunc="sum", fill_value=0, margins=True, margins_name="Grand Total"
    )
    weekly_pivot.columns = [c.strftime("%Y-%m-%d") if hasattr(c, "strftime") else c
                             for c in weekly_pivot.columns]

    quality_log_df = pd.DataFrame({"Pipeline Run Log": log_lines})

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        sku_region_pivot.to_excel(writer, sheet_name="Pivot - SKU x Region")
        weekly_pivot.to_excel(writer, sheet_name="Pivot - Weekly Trend")
        df.drop(columns=["Week Starting"]).to_excel(writer, sheet_name="Consolidated Data", index=False)
        quality_log_df.to_excel(writer, sheet_name="Data Quality Log", index=False)

    _style_workbook(out_path)


def _style_workbook(path):
    """Light formatting pass: bold header row, professional font, sensible widths."""
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", size=10)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for col_cells in ws.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 22)
        ws.freeze_panes = "A2"
    wb.save(path)


def main():
    log(f"Pipeline run started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log(f"Scanning {RAW_DIR} for raw planner exports...")

    files = sorted(glob.glob(f"{RAW_DIR}/*.csv") + glob.glob(f"{RAW_DIR}/*.xlsx"))
    if not files:
        log("No raw files found. Nothing to do.")
        return

    frames = []
    for path in files:
        log(f"Reading {os.path.basename(path)}")
        df = read_raw_file(path)
        if df is not None:
            log(f"  {len(df)} rows read")
            frames.append(df)

    if not frames:
        log("No usable files found. Aborting.")
        return

    combined = pd.concat(frames, ignore_index=True)
    log(f"\nCombined raw rows from all files: {len(combined)}")
    log("Cleaning consolidated dataset...")
    cleaned = clean_dataset(combined)

    consolidated_path = f"{OUTPUT_DIR}/consolidated_data.csv"
    cleaned.to_csv(consolidated_path, index=False)
    log(f"\nSaved cleaned dataset -> {consolidated_path}")

    report_path = f"{OUTPUT_DIR}/Weekly_Supply_Chain_Summary_Report.xlsx"
    build_summary_report(cleaned.copy(), report_path)
    log(f"Saved summary pivot report -> {report_path}")

    log_path = f"{OUTPUT_DIR}/pipeline_log.txt"
    with open(log_path, "w") as f:
        f.write("\n".join(log_lines))
    print(f"\nLog written -> {log_path}")


if __name__ == "__main__":
    main()
